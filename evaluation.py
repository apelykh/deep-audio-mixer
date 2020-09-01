import os
import soundfile as sf
# import librosa
import torch
import pickle
import numpy as np
import pyloudnorm as pyln
from openpyxl import Workbook
from statistics import mean
from collections import OrderedDict

from inference_utils import mix_song_smooth
from data.dataset import MultitrackAudioDataset
from data.dataset_utils import load_tracks_musdb18
from models.model_scalar_1s import MixingModelScalar1s
from models.model_scalar_2s import MixingModelScalar2s
from models.baselines.mean_loudness_model import MeanLoudnessModel
from models.baselines.random_model import RandomModel


class LoudnessEvaluator:
    def __init__(self, dataset, d_mean_loudness, mix_model, sr=44100, seed=None):
        if seed:
            np.random.seed(seed)

        self.sr = sr
        self.d = dataset
        self.mix_model = mix_model
        self.mean_loudness_model = MeanLoudnessModel(d_mean_loudness)
        self.random_model = RandomModel()

        self.meter = pyln.Meter(sr)
        self.keys = ('bass', 'drums', 'vocals', 'other')

        self.results_dir = './experiment'
        if not os.path.exists(self.results_dir):
            os.makedirs(self.results_dir)

    def evaluate_loudness(self, tracks: dict) -> list:
        per_track_loudness = [self.meter.integrated_loudness(tracks[track_name].T)
                              for track_name in self.keys]
        avg_loudness = mean(per_track_loudness)
        per_track_loudness_norm = [track_loudness - avg_loudness
                                   for track_loudness in per_track_loudness]

        return per_track_loudness_norm

    @staticmethod
    def _calculate_diff_between_loudness_dicts(l_dict1: OrderedDict, l_dict2: OrderedDict):
        a1 = np.array(list(l_dict1.values()))
        a2 = np.array(list(l_dict2.values()))
        diffs = np.abs(a1 - a2)
        return float(np.mean(diffs))

    def _sum_and_evaluate_tracks(self, track_dict, reference_dict, song_name, identifier, write_to_disk=True):
        print('[.] {}'.format(identifier.upper()))

        if write_to_disk:
            track_arr = np.array(list(track_dict.values()))
            track_sum = np.sum(track_arr, axis=0)

            # track_sum = librosa.util.normalize(track_sum, axis=1)
            loudness = self.meter.integrated_loudness(track_sum.T)
            loudness_norm_sum = pyln.normalize.loudness(track_sum.T, loudness, -20.0)
            sf.write(os.path.join(self.results_dir, '{}_{}.wav'.format(song_name, identifier)),
                     loudness_norm_sum, self.sr)

        per_track_loudness = self.evaluate_loudness(track_dict)
        loudness_dict = OrderedDict(zip(self.keys, per_track_loudness))

        if reference_dict:
            error = LoudnessEvaluator. \
                _calculate_diff_between_loudness_dicts(loudness_dict, reference_dict)
            return loudness_dict, error
        return loudness_dict, None

    def process_song(self, base_dir: str, song_name: str, n_random_samples: int = 5,
                     write_wavs_to_disk=False) -> dict:
        stats = {
            'song_name': song_name
        }

        # loading reference
        loaded_tracks = load_tracks_musdb18(os.path.join(base_dir, 'manual_gain_mixes'), song_name,
                                            tracklist=('bass', 'drums', 'vocals', 'other'))
        reference, _ = self._sum_and_evaluate_tracks(loaded_tracks, None, song_name,
                                                     identifier='reference',
                                                     write_to_disk=write_wavs_to_disk)

        loaded_tracks = load_tracks_musdb18(os.path.join(base_dir, 'test'), song_name,
                                            tracklist=('bass', 'drums', 'vocals', 'other'))
        _, stats['sum_error'] = self._sum_and_evaluate_tracks(loaded_tracks, reference, song_name,
                                                              identifier='sum',
                                                              write_to_disk=write_wavs_to_disk)

        # each multitrack is normalized to the mean loudness of the corresponding track from train set
        loudnorm_tracks = self.mean_loudness_model.forward(loaded_tracks)
        _, stats['loudnorm_error'] = self._sum_and_evaluate_tracks(loudnorm_tracks, reference, song_name,
                                                                   identifier='loudnorm',
                                                                   write_to_disk=write_wavs_to_disk)
        # mixed by the model
        mixed_tracks, _, _ = mix_song_smooth(self.d, self.mix_model, loaded_tracks, chunk_length=2)
        _, stats['mix_error'] = self._sum_and_evaluate_tracks(mixed_tracks, reference, song_name,
                                                              identifier='mix',
                                                              write_to_disk=write_wavs_to_disk)
        random_errors = []
        for exp_i in range(n_random_samples):
            # sum of multitracks with random gains
            random_mix_tracks = self.random_model.forward(loaded_tracks)
            _, random_error = self._sum_and_evaluate_tracks(random_mix_tracks, reference, song_name,
                                                            identifier='random_{}'.format(exp_i),
                                                            write_to_disk=write_wavs_to_disk)
            random_errors.append(random_error)
        stats['random_error'] = mean(random_errors)

        return stats

    def process_songlist(self, base_dir, songlist, write_to_disk=False):
        book = Workbook()
        sheet = book.active

        keys = ['song_name', 'sum_error', 'random_error', 'loudnorm_error', 'mix_error']
        for key_i, key in enumerate(keys):
            sheet.cell(column=key_i + 1, row=1, value=key)

        errors = {key: [] for key in keys[1:]}

        for i, song_name in enumerate(songlist):
            print('{}/{}: {}'.format(i + 1, len(songlist), song_name))
            stats = self.process_song(base_dir, song_name, write_wavs_to_disk=write_to_disk)

            for key_i, key in enumerate(keys):
                value = stats[key] if key == 'song_name' else '{:.4f}'.format(stats[key])
                sheet.cell(column=key_i + 1, row=i + 2, value=value)

            for key in keys[1:]:
                errors[key].append(stats[key])

        final_row = len(songlist) + 2
        for key_i, key in enumerate(keys):
            value = 'Mean' if key == 'song_name' else '{:.2f}'.format(mean(errors[key]))
            sheet.cell(column=key_i + 1, row=final_row, value=value)

        book.save('./stats.xlsx')


if __name__ == '__main__':
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print('Torch version: ', torch.__version__)
    print('Device: ', device)

    base_path = '/media/apelykh/bottomless-pit/datasets/mixing/MedleyDB/Audio'
    chunk_length = 2

    # 2s 2 datasets
    train_songlist = ['Actions - One Minute Smile', 'ANiMAL - Clinic A', "Actions - Devil's Words", 'Young Griffo - Blood To Bone',
         'Jokers, Jacks & Kings - Sea Of Leaves', 'Auctioneer_OurFutureFaces', 'Grants_PunchDrunk', 'Giselle - Moss',
         'The Long Wait - Back Home To Blue', 'Remember December - C U Next Time', 'Johnny Lokke - Whisper To A Scream',
         "The Wrong'Uns - Rothko", 'Chris Durban - Celebrate', 'St Vitus - Word Gets Around',
         'AClassicEducation_NightOwl', 'Grants - PunchDrunk', 'InvisibleFamiliars_DisturbingWildlife',
         'Skelpolu - Human Mistakes', 'PurlingHiss_Lolita', 'Jay Menon - Through My Eyes', 'James May - On The Line',
         'FamilyBand_Again', 'StevenClark_Bounty', 'HeladoNegro_MitadDelMundo', 'DreamersOfTheGhetto_HeavyLove',
         'Fergessen - The Wind', 'James May - All Souls Moon', 'Bill Chudziak - Children Of No-one',
         'BigTroubles_Phantom', 'Dark Ride - Burning Bridges', 'Fergessen - Nos Palpitants',
         'North To Alaska - All The Same', 'PortStWillow_StayEven', 'SweetLights_YouLetMeDown', 'ANiMAL - Easy Tiger',
         'Leaf - Summerghost', 'HezekiahJones_BorrowedHeart', 'Hollow Ground - Left Blind',
         'Johnny Lokke - Promises & Lies', 'Atlantis Bound - It Was My Fault For Waiting',
         'Voelund - Comfort Lives In Belief', 'Swinging Steaks - Lost My Way', 'Young Griffo - Facade',
         'Titanium - Haunted Age', 'Traffic Experiment - Once More (With Feeling)',
         "Phre The Eon - Everybody's Falling Apart", 'Black Bloc - If You Want Success',
         'Angela Thomas Wade - Milk Cow Blues', 'Flags - 54', 'Patrick Talbot - A Reason To Leave',
         'TheDistricts_Vermont', 'Leaf - Wicked', 'Creepoid_OldTree', 'HopAlong_SisterCities', 'AvaLuna_Waterduct',
         'SecretMountains_HighHorse', 'Drumtracks - Ghost Bitch', 'Cnoc An Tursa - Bannockburn',
         'Patrick Talbot - Set Me Free', 'Triviul - Angelsaint', 'FacesOnFilm_WaitingForGa', 'Triviul - Dorothy',
         'Skelpolu - Together Alone', 'Actions - South Of The Water']

    d = MultitrackAudioDataset(
        base_path,
        songlist=train_songlist,
        chunk_length=chunk_length,
        normalize=False,
        compute_features=True
    )

    if not os.path.exists('./mean_loudness.pkl'):
        mean_loudness = d.compute_mean_loudness()
        with open("mean_loudness.pkl", "wb") as f:
            pickle.dump(mean_loudness, f)
    else:
        with open("mean_loudness.pkl", "rb") as f:
            mean_loudness = pickle.load(f)

    model = MixingModelScalar2s().to(device)
    # 2.4694 mean, but some songs are good
    # weights = './saved_models/from_server/21-08-2020-10:49_training_4masks_unnorm_1s_medleydb+musdb_train_30_epochs_110_val_loss/scalar1s_23_neg_train_loss=-109.7035.pt'

    # 2.1715 mean!
    weights = './saved_models/from_server/21-08-2020-16:44_training_4masks_unnorm_2s_medleydb+musdb_train_26_epochs_168.41_val_loss/best_scalar2s_19_neg_train_loss=-168.4133.pt'
    model.load_state_dict(torch.load(weights, map_location=device))

    base_dir = '/media/apelykh/bottomless-pit/datasets/mixing/MUSDB18HQ'
    test_songlist = [
        "Arise - Run Run Run",
        "BKS - Bulldozer",
        "BKS - Too Much",
        "Bobby Nobody - Stitch Up",
        "Cristina Vane - So Easy",
        "Enda Reilly - Cur An Long Ag Seol",
        "Forkupines - Semantics",
        "James Elder & Mark M Thompson - The English Actor",
        "Nerve 9 - Pray For The Rain",
        "Raft Monk - Tiring",
        "Signe Jakobsen - What Have You Done To Me",
        "Speak Softly - Broken Man",
        "The Doppler Shift - Atrophy",
        "Timboz - Pony",
        "Zeno - Signs"
    ]

    l_eval = LoudnessEvaluator(d, mean_loudness, model)
    l_eval.process_songlist(base_dir, test_songlist)
